import openpyxl
import xlrd
team=['ENG','SA','IND','AUS','NZ','PAK','BAN','SL','AF','WI']
xwb1=xlrd.open_workbook("D:\Important Data\Book1.xlsx")
xsheet1=xwb1.sheet_by_index(0)
norow1=xsheet1.nrows
nocol1=xsheet1.ncols
owb1=openpyxl.load_workbook("D:\Important Data\Book1.xlsx")
xsheet2=xwb1.sheet_by_index(1)
osheet2=owb1.get_sheet_by_name("Sheet2")
norow2=xsheet2.nrows
nocol2=xsheet2.ncols
winteamlist=[]
print(norow1)
print(nocol1)
print(xsheet1.cell_value(0,0))
print(osheet2.title)
# osheet2.cell(row=3,column=3).value=7
k=0
for i in range(7,nocol1):
    for j in range(1,norow1-3):
        k+=1
        if xsheet1.cell_value(j,4)==xsheet1.cell_value(j,i):
            winteamlist.append(xsheet1.cell_value(j,i))
        else:
            winteamlist.append('')

n=int(k/45)
print(n)
m=0
print(nocol2)
print(len(winteamlist))
for i in range(3,nocol2+n+1):
    start=m*45
    end=45*(m+1)
    print(start,end)
    winteam=winteamlist[start:end]
    for j in range(3,norow2+1):
        pts=2*(winteam.count(team[j-3]))
        osheet2.cell(row=j,column=i).value=pts
        print(j,i,osheet2.cell(row=j,column=i).value)

    m+=1


owb1.save("D:\Important Data\Book1.xlsx")

